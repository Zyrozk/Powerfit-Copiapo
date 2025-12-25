from django.shortcuts import render
from .serializer import agendarClaseSerializer, datosFisicosSerializer, registroSerializer, coachSerializer, membresiaSerializer, nuevoCoachSerializer, claseProgramadaSerializer, pagoSerializer
from .models import AgendarClase, DatosFisicos, Registro, Coach, Membresia, NuevoCoach, ClaseProgramada, Pago, ReservaClase, Asistencia
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
from rest_framework.parsers import MultiPartParser, FormParser
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from functools import wraps
from datetime import timedelta, date, datetime
from django.db.models import Count, Sum, Q
from django.db import models
from .models import Membresia, Asistencia, Pago
import qrcode
from django.shortcuts import redirect
from django.core.mail import send_mail


# Create your views here.

def solo_admin(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        rol = request.headers.get("Rol")
        # Solo admin puede ejecutar estas vistas
        if rol != "admin":
            return Response({"error": "No autorizado. Solo admin puede realizar esta acción."}, status=403)
        return view_func(request, *args, **kwargs)
    return wrapper


@api_view(['GET', 'POST'])
def agendarClase_list(request):
    if request.method == 'GET':
        clases = AgendarClase.objects.all()
        serializer = agendarClaseSerializer(clases, many=True)
        return Response(serializer.data)
    
    if request.method == 'POST':
        serializer = agendarClaseSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

@api_view(['GET', 'POST'])
def datosFisicos_list(request):
    if request.method == 'GET':
        datos = DatosFisicos.objects.all()
        serializer = datosFisicosSerializer(datos, many=True)
        return Response(serializer.data)
    
    if request.method == 'POST':
        serializer = datosFisicosSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def datosFisicos_usuario(request, user_id):
    datos = DatosFisicos.objects.filter(usuario_id=user_id).order_by('-fecha')
    serializer = datosFisicosSerializer(datos, many=True)
    return Response(serializer.data)


@api_view(['GET', 'POST', 'PATCH'])
def registro_list(request):
    if request.method == 'GET':
        registros = Registro.objects.all()
        serializer = registroSerializer(registros, many=True)
        return Response(serializer.data)
    
    if request.method == 'POST':
        serializer = registroSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    if request.method == 'PATCH':
        user_id = request.data.get("id")
        if not user_id:
            return Response({"error": "Debe enviar id del usuario"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            registro = Registro.objects.get(id=user_id)
        except Registro.DoesNotExist:
            return Response({"error": "Usuario no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        serializer = registroSerializer(registro, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
@api_view(['GET', 'POST', 'PATCH'])
def coach_list(request):
    if request.method == 'GET':
        entrenador = Coach.objects.all()
        serializer = coachSerializer(entrenador, many=True)
        return Response(serializer.data)
    
    if request.method == 'POST':
        serializer = coachSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    if request.method == 'PATCH':
        serializer = coachSerializer(Coach, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)
    
    
@api_view(['POST'])
def membresia_list(request):
    usuario_id = request.data.get("usuario")

    usuario = Registro.objects.get(id=usuario_id)

    activa = Membresia.objects.filter(usuario=usuario, is_active=True).first()
    if activa:
        return Response({"error": "Ya tienes una membresía activa"}, status=400)

    serializer = membresiaSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)

@api_view(['POST'])
def comprar_membresia_list(request):
    # 1. Recibir datos (incluyendo archivos)
    usuario_id = request.data.get("usuario")

    try:
        usuario_instancia = Registro.objects.get(id=usuario_id)
    except Registro.DoesNotExist:
        return Response({"error": "Usuario no encontrado"}, status=404)

    # 2. Verificar si el usuario ya tiene una membresía activa
    membresia_activa = Membresia.objects.filter(
        usuario=usuario_instancia,
        is_active=True,
        end_date__gte=timezone.now().date()
    ).exists()

    if membresia_activa:
        return Response({"error": "El usuario ya tiene una membresía activa. No puede adquirir otra hasta que expire."}, status=400)

    # 3. Crear membresía en estado Pendiente
    nueva_membresia = Membresia.objects.create(
        usuario=usuario_instancia,
        plan_nombre=request.data.get("plan_nombre"),
        plan_clases=request.data.get("plan_clases"),
        plan_precio=request.data.get("plan_precio"),
        comprobante=request.FILES.get('comprobante'),
        estado='Pendiente',
        is_active=False
    )

    # 4. 🔥 Crear registro de pago automático
    Pago.objects.create(
        usuario=usuario_instancia,
        tipo="membresía",
        monto=request.data.get("plan_precio"),
        metodo="transferencia",   # Puedes cambiarlo si lo deseas
        estado="Pendiente"
    )

    return Response({"mensaje": "Solicitud enviada"}, status=201)


@api_view(['GET'])
def membresia_activa(request, user_id):
    try:
        usuario = Registro.objects.get(id=user_id)
    except Registro.DoesNotExist:
        return Response({"activa": False})

    membresia = Membresia.objects.filter(usuario=usuario, is_active=True).first()

    if not membresia:
        return Response({"activa": False})

    return Response({
        "activa": True,
        "plan_nombre": membresia.plan_nombre,
        "plan_clases": membresia.plan_clases,
        "plan_precio": membresia.plan_precio,
        "start_date": membresia.start_date,
        "end_date": membresia.end_date,
    })

# NUEVA VISTA: Para que el Coach apruebe
@api_view(['POST'])
def aprobar_membresia(request, membresia_id):
    try:
        membresia = Membresia.objects.get(id=membresia_id)
    except Membresia.DoesNotExist:
        return Response({"error": "Membresía no encontrada"}, status=404)

    # 1. Activar la membresía
    membresia.estado = 'Activa'
    membresia.is_active = True
    membresia.save()

    # 2. Buscar el pago pendiente asociado a esta membresía
    pago = Pago.objects.filter(
        usuario=membresia.usuario,
        tipo="membresía",
        estado="Pendiente"
    ).first()

    # 3. Actualizar el pago si existe
    if pago:
        pago.estado = "Realizado"
        pago.save()

    return Response({"mensaje": "Membresía aprobada y pago confirmado"})


# NUEVA VISTA: Listar todas las membresías (para el panel del coach)
@api_view(['GET'])
def listar_todas_membresias(request):
    membresias = Membresia.objects.all().order_by('-start_date')
    data = []
    for m in membresias:
        data.append({
            "id": m.id,
            "user": f"{m.usuario.nombre} {m.usuario.apellido}",
            "plan_nombre": m.plan_nombre,
            "plan_precio": m.plan_precio,
            "plan_clases": m.plan_clases,
            "start_date": m.start_date,
            "status": m.estado,
            "comprobante": m.comprobante.url if m.comprobante else None
        })
    return Response(data)


@api_view(['GET'])
def mis_reservas(request, user_id):
    """Devuelve las reservas (historial de clases) de un usuario."""
    reservas = ReservaClase.objects.filter(usuario_id=user_id).order_by('-fecha_reserva')

    data = []
    for r in reservas:
        data.append({
            'id': r.id,
            'clase_id': r.clase.id,
            'tipo': r.clase.tipo,
            'dia': r.clase.dia,
            'fecha': r.clase.fecha.strftime('%Y-%m-%d') if r.clase.fecha else None,
            'horario': r.clase.horario,
            'reserved_at': r.fecha_reserva.strftime('%Y-%m-%d %H:%M:%S'),
            'estado': r.estado,
        })

    return Response(data)


@api_view(['GET'])
def mis_pagos(request, user_id):
    """Devuelve los pagos realizados/registrados por un usuario."""
    pagos = Pago.objects.filter(usuario_id=user_id).order_by('-fecha', '-hora')

    data = []
    for p in pagos:
        data.append({
            'id': p.id,
            'usuario': p.usuario.id,
            'usuario_nombre': getattr(p.usuario, 'nombre', None),
            'monto': p.monto,
            'tipo': p.tipo,
            'metodo': p.metodo,
            'fecha': p.fecha.strftime('%Y-%m-%d'),
            'hora': p.hora.strftime('%H:%M:%S') if p.hora else None,
            'estado': p.estado,
        })

    return Response(data)


@api_view(['POST'])
def login_usuario(request):
    correo = request.data.get("correo")
    password = request.data.get("password")

    # 1️⃣ ¿Es Gabriela o un coach?
    user = authenticate(username=correo, password=password)

    if user:
        if user.is_superuser:
            return Response({
                "id": user.id,
                "nombre": user.first_name,
                "apellido": user.last_name,
                "correo": user.username,
                "rol": "admin"
            })
        else:
            return Response({
                "id": user.id,
                "nombre": user.first_name,
                "apellido": user.last_name,
                "correo": user.username,
                "rol": "coach"
            })

    try:
        u = Registro.objects.get(correo=correo, password=password)
        return Response({
            "id": u.id,
            "nombre": u.nombre,
            "apellido": u.apellido,
            "correo": u.correo,
            "rol": "usuario"
        })
    except Registro.DoesNotExist:
        pass

    return Response({"error": "Credenciales incorrectas"}, status=400)

@api_view(['DELETE'])
def eliminar_membresia(request, membresia_id):
    try:
        membresia = Membresia.objects.get(id=membresia_id)
        membresia.delete()
        return Response({"mensaje": "Membresía eliminada"}, status=status.HTTP_204_NO_CONTENT)
    except Membresia.DoesNotExist:
        return Response({"error": "Membresía no encontrada"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET', 'POST'])
def coach_list(request):

    # 📌 GET — Cualquier coach puede ver la lista
    if request.method == 'GET':
        coaches = NuevoCoach.objects.all()
        serializer = nuevoCoachSerializer(coaches, many=True)
        return Response(serializer.data)

    # 📌 POST — SOLO admin puede crear coaches
    if request.method == 'POST':

        nombre = request.data.get("nombre")
        apellido = request.data.get("apellido")
        correo = request.data.get("correo")
        password = request.data.get("password")
        bibliografia = request.data.get("bibliografia")
        especialidad = request.data.get("especialidad")

        # 1️⃣ Crear registro en NuevoCoach
        coach = NuevoCoach.objects.create(
            nombre=nombre,
            apellido=apellido,
            correo=correo,
            password=password,  # referencia, login usa Django Auth
            bibliografia=bibliografia,
            especialidad=especialidad
        )

        # 2️⃣ Crear usuario Django Auth
        user = User.objects.create_user(
            username=correo,
            email=correo,
            password=password,
            first_name=nombre,
            last_name=apellido,
            is_staff=False,       # NO admin
            is_superuser=False    # NO superadmin
        )

        return Response({"mensaje": "Coach creado correctamente"}, status=201)


    
@api_view(['GET', 'PUT', 'DELETE'])
def coach_detail(request, coach_id):

    try:
        coach = NuevoCoach.objects.get(id=coach_id)
    except NuevoCoach.DoesNotExist:
        return Response({"error": "Coach no encontrado"}, status=404)

    if request.method == 'GET':
        serializer = nuevoCoachSerializer(coach)
        return Response(serializer.data)

    if request.method == 'PUT':
        serializer = nuevoCoachSerializer(coach, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    if request.method == 'DELETE':
        coach.delete()
        return Response({"mensaje": "Coach eliminado"}, status=204)
    
DIAS_ORDEN = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]

@api_view(['GET', 'POST'])
def horario_list(request):

    if request.method == 'GET':
        clases = ClaseProgramada.objects.all()

        # Ordenar primero por día (usando DIAS_ORDEN), luego por horario
        clases = sorted(
            clases,
            key=lambda c: (DIAS_ORDEN.index(c.dia), c.horario)
        )

        serializer = claseProgramadaSerializer(clases, many=True)
        return Response(serializer.data)

    if request.method == 'POST':
        serializer = claseProgramadaSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)


@api_view(['DELETE'])
def horario_delete(request, pk):
    try:
        clase = ClaseProgramada.objects.get(id=pk)
    except ClaseProgramada.DoesNotExist:
        return Response({"error": "Clase no encontrada"}, status=404)

    clase.delete()
    return Response({"mensaje": "Clase eliminada"}, status=204)


@api_view(['GET'])
def pagos_list(request):
    fecha_inicio = request.GET.get("inicio")
    fecha_fin = request.GET.get("fin")

    # Mostrar SOLO pagos realizados
    pagos = Pago.objects.filter(estado="Realizado").order_by("-fecha", "-hora")

    # Filtro por fecha
    if fecha_inicio and fecha_fin:
        pagos = pagos.filter(fecha__range=[fecha_inicio, fecha_fin])

    serializer = pagoSerializer(pagos, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def exportar_pagos(request):
    fecha_inicio = request.GET.get("inicio")
    fecha_fin = request.GET.get("fin")

    # Mostrar solo pagos realizados
    pagos = Pago.objects.filter(estado="Realizado")

    if fecha_inicio and fecha_fin:
        pagos = pagos.filter(fecha__range=[fecha_inicio, fecha_fin])

    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"

    # Encabezados
    headers = ["Usuario", "Tipo", "Monto", "Método", "Fecha", "Hora", "Estado"]
    ws.append(headers)

    # Encabezados en negrita
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Datos
    for p in pagos:
        ws.append([
            f"{p.usuario.nombre} {p.usuario.apellido}",  # Nombre completo
            p.tipo,
            p.monto,
            p.metodo,
            p.fecha.strftime("%d-%m-%Y"),
            p.hora.strftime("%H:%M"),
            p.estado
        ])

    from openpyxl.utils import get_column_letter

    for column_cells in ws.columns:
        max_length = 0

        for cell in column_cells:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass

        adjusted_width = max_length + 2
        column_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    response["Content-Disposition"] = 'attachment; filename="pagos.xlsx"'
    return response

@api_view(['POST'])
def reservar_clase(request):
    clase_id = request.data.get("clase")
    usuario_id = request.data.get("usuario")
    usar_membresia = request.data.get("usar_membresia")
    pagar_efectivo = request.data.get("pagar_efectivo")

    if not clase_id or not usuario_id:
        return Response({"error": "Faltan datos"}, status=400)

    # Obtener clase y usuario
    try:
        clase = ClaseProgramada.objects.get(id=clase_id)
        usuario = Registro.objects.get(id=usuario_id)
    except (ClaseProgramada.DoesNotExist, Registro.DoesNotExist):
        return Response({"error": "Clase o usuario no encontrado"}, status=404)

    # No permitir sobrecupo
    if clase.cupos_disponibles <= 0:
        return Response({"error": "No quedan cupos disponibles"}, status=400)

    # Evitar reservas duplicadas
    if ReservaClase.objects.filter(clase=clase, usuario=usuario).exists():
        return Response({"error": "Ya estás inscrito en esta clase"}, status=400)
    # Si se indica usar membresía, verificar y descontar
    if str(usar_membresia).lower() in ["1", "true", "yes"]:
        membresia = Membresia.objects.filter(usuario=usuario, is_active=True).first()
        if not membresia:
            return Response({"error": "No tienes una membresía activa"}, status=400)

        if membresia.plan_clases <= 0:
            return Response({"error": "Tu membresía no tiene clases disponibles"}, status=400)

        # Crear reserva
        ReservaClase.objects.create(
            clase=clase,
            usuario=usuario,
            estado="Reservado"
        )

        # Descontar cupo
        clase.cupos_disponibles -= 1
        clase.save()

        # Descontar una clase de la membresía
        membresia.plan_clases -= 1
        if membresia.plan_clases <= 0:
            membresia.is_active = False
        membresia.save()

        return Response({"mensaje": "Clase reservada usando membresía"}, status=201)

    # Si se paga en efectivo (pago inmediato para la clase)
    if str(pagar_efectivo).lower() in ["1", "true", "yes"]:
        # Crear pago en efectivo y marcarlo como realizado
        try:
            pago = Pago.objects.create(
                usuario=usuario,
                tipo="clase",
                monto=3500,
                metodo="efectivo",
                estado="Realizado"
            )
        except Exception:
            return Response({"error": "No se pudo procesar el pago"}, status=500)

        # Crear reserva
        ReservaClase.objects.create(
            clase=clase,
            usuario=usuario,
            estado="Reservado"
        )

        # Descontar cupo
        clase.cupos_disponibles -= 1
        clase.save()

        return Response({"mensaje": "Clase reservada y pagada en efectivo", "pago_id": pago.id}, status=201)

    # Si no se especifica método, crear reserva normal (sin pago)
    ReservaClase.objects.create(
        clase=clase,
        usuario=usuario,
        estado="Reservado"
    )

    # Descontar cupo
    clase.cupos_disponibles -= 1
    clase.save()

    return Response({"mensaje": "Clase reservada correctamente"}, status=201)



@api_view(['GET'])
def clases_de_dia(request):
    # Intentamos obtener las clases usando la fecha local del servidor (más fiable que comparar nombres)
    hoy = timezone.localdate()

    # Buscar por fecha (si las clases se guardan con fecha exacta)
    clases = ClaseProgramada.objects.filter(fecha=hoy)

    # Si no hay clases por fecha, hacemos fallback a la búsqueda por nombre de día
    if not clases.exists():
        dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        # weekday(): 0=Lunes ... 6=Domingo
        try:
            dia_actual = dias[timezone.now().weekday()]
        except Exception:
            dia_actual = None
        if dia_actual:
            clases = ClaseProgramada.objects.filter(dia=dia_actual)

    serializer = claseProgramadaSerializer(clases, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def asistentes_clase(request, clase_id):
    try:
        clase = ClaseProgramada.objects.get(id=clase_id)
    except ClaseProgramada.DoesNotExist:
        return Response({"error": "Clase no encontrada"}, status=404)

    # Use ReservaClase (las reservas reales) para listar los inscritos.
    # El endpoint `reservar_clase` crea objetos ReservaClase, no Asistencia,
    # por lo que debemos consultarlas aquí para mostrar todos los alumnos
    # que se han inscrito en la clase.
    reservas = ReservaClase.objects.filter(clase=clase)

    data = [
        {
            "id": r.id,
            "usuario_id": r.usuario.id,
            "nombre": r.usuario.nombre,
            "apellido": r.usuario.apellido,
            "nombre_completo": f"{r.usuario.nombre} {r.usuario.apellido}",
            # Estado de presencia se almacena en ReservaClase.estado (Presente/Ausente/Pendiente)
            "presente": True if r.estado == "Presente" else False
        }
        for r in reservas
    ]

    return Response(data)



@api_view(['PATCH'])
def marcar_presente(request, reserva_id):
    try:
        reserva = ReservaClase.objects.get(id=reserva_id)
    except ReservaClase.DoesNotExist:
        return Response({"error": "Reserva no encontrada"}, status=404)

    reserva.estado = "Presente"
    reserva.save()

    return Response({"mensaje": "Asistencia registrada"})


def get_start_date(range):
    now = timezone.now()

    if range == "day":
        return now - timedelta(days=1)
    if range == "week":
        return now - timedelta(weeks=1)
    if range == "month":
        return now - timedelta(days=30)
    if range == "6months":
        return now - timedelta(days=180)
    if range == "year":
        return now - timedelta(days=365)

    return now - timedelta(days=30)


@api_view(["GET"])
def estadisticas_membresias(request):
    range = request.GET.get("range", "month")
    start_date = get_start_date(range)

    data = (
        Membresia.objects
        .filter(start_date__gte=start_date)
        .values("start_date__month")
        .annotate(total=Count("id"))
        .order_by("start_date__month")
    )

    return Response(data)


@api_view(["GET"])
def estadisticas_asistencia(request):
    range = request.GET.get("range", "month")
    start_date = get_start_date(range)

    data = (
        Asistencia.objects
        .filter(fecha_reserva__gte=start_date)
        .values("clase__tipo")
        .annotate(
            total=Count("id"),
            presentes=Count("id", filter=Q(presente=True))
        )
    )

    result = []
    for d in data:
        porcentaje = (d["presentes"] / d["total"]) * 100 if d["total"] > 0 else 0
        result.append({
            "clase": d["clase__tipo"],
            "porcentaje": round(porcentaje, 1)
        })

    return Response(result)


@api_view(["GET"])
def estadisticas_pagos(request):
    range = request.GET.get("range", "month")
    start_date = get_start_date(range)

    data = (
        Pago.objects
        .filter(fecha__gte=start_date, estado="Aprobado")
        .values("fecha__month")
        .annotate(total=Sum("monto"))
        .order_by("fecha__month")
    )

    return Response(data)


@api_view(["GET"])
def clases_hoy(request):
    hoy = date.today()

    clases = ClaseProgramada.objects.filter(fecha=hoy)

    data = [
        {
            "id": c.id,
            "tipo": c.tipo,
            "horario": c.horario,
            "fecha": c.fecha.strftime("%d-%m-%Y"),
        }
        for c in clases
    ]

    return Response(data)


@api_view(["POST"])
def confirmar_asistencia(request):
    usuario_id = request.data.get("usuario_id")
    clase_id = request.data.get("clase_id")

    try:
        reserva = ReservaClase.objects.get(
            usuario_id=usuario_id,
            clase_id=clase_id,
            estado="Pendiente"
        )

        reserva.estado = "Presente"
        reserva.save()

        return Response(
            {"message": "Asistencia confirmada"},
            status=status.HTTP_200_OK
        )

    except ReservaClase.DoesNotExist:
        return Response(
            {"error": "Reserva no válida"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
def procesar_ausencias():
    hoy = timezone.now().date()

    reservas = ReservaClase.objects.filter(
        estado="Pendiente",
        clase__fecha__lt=hoy
    )

    for reserva in reservas:
        reserva.estado = "Ausente"
        reserva.save()

        membresia = Membresia.objects.filter(
            usuario=reserva.usuario,
            is_active=True
        ).first()

        if membresia:
            membresia.clases_disponibles += 1
            membresia.save()

@api_view(["GET"])
def asistentes_por_clase(request, clase_id):
    reservas = ReservaClase.objects.filter(
        clase_id=clase_id,
        estado="Presente"
    )

    data = [
        {
            "nombre": r.usuario.nombre,
            "apellido": r.usuario.apellido
        }
        for r in reservas
    ]

    return Response(data)


@api_view(["GET"])
def generar_qr(request, clase_id):
    url = f"http://localhost:5173/asistencia/confirmar/{clase_id}"

    img = qrcode.make(url)
    response = HttpResponse(content_type="image/png")
    img.save(response, "PNG")
    return response


@api_view(["GET"])
def historial_uso(request, user_id):
    reservas = ReservaClase.objects.filter(
        usuario_id=user_id
    ).exclude(estado="Pendiente")

    data = [
        {
            "tipo": r.clase.tipo,
            "fecha": r.clase.fecha.strftime("%d-%m-%Y"),
            "horario": r.clase.horario,
            "estado": r.estado
        }
        for r in reservas
    ]

    return Response(data)

def clases_por_fecha(request):
    fecha = request.GET.get("fecha")

    if not fecha:
        return JsonResponse([], safe=False)

    try:
        fecha_obj = datetime.strptime(fecha, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse([], safe=False)

    clases = ClaseProgramada.objects.filter(fecha=fecha_obj)

    data = []
    for c in clases:
        data.append({
            "id": c.id,
            "tipo": c.tipo,
            "horario": c.horario,
            "fecha": c.fecha.strftime("%Y-%m-%d"),
        })

    return JsonResponse(data, safe=False)

def registrar_asistencia_qr(request, clase_id):
    if not request.user.is_authenticated:
        return redirect("/login")

    reserva = ReservaClase.objects.filter(
        usuario=request.user,
        clase_id=clase_id
    ).first()

    if not reserva:
        return JsonResponse({"error": "No tienes reserva"}, status=403)

    if reserva.estado == "Presente":
        return JsonResponse({"message": "Ya registrada"})

    reserva.estado = "Presente"
    reserva.save()

    return JsonResponse({"ok": True})


def revisar_membresias():
    hoy = timezone.now().date()
    aviso = hoy + timedelta(days=3)

    membresias = Membresia.objects.filter(is_active=True)

    for m in membresias:
        # 🔔 Por vencer
        if m.end_date == aviso:
            send_mail(
                subject="Tu membresía está por vencer",
                message=f"Hola {m.usuario.nombre}, tu membresía vence el {m.end_date}.",
                from_email="powerfit@gym.cl",
                recipient_list=[m.usuario.correo],
            )

        # ⛔ Vencida
        if m.end_date < hoy:
            send_mail(
                subject="Tu membresía ha vencido",
                message=f"Hola {m.usuario.nombre}, tu membresía venció el {m.end_date}.",
                from_email="powerfit@gym.cl",
                recipient_list=[m.usuario.correo],
            )
            m.is_active = False
            m.save()
